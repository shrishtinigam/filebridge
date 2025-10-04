import subprocess
from pathlib import Path
import tempfile
from openpyxl import load_workbook, Workbook
from openpyxl.utils import range_boundaries, get_column_letter
from typing import Union, List, Tuple, Dict


class ExcelExporter:
    """
    Export Excel ranges or charts to PDF/PNG using LibreOffice headless.
    """

    def __init__(self, workbook: Union[str, Workbook]):
        """
        Accept a path or an openpyxl Workbook.
        """
        if isinstance(workbook, str):
            self.file_path = Path(workbook).absolute()
        elif isinstance(workbook, Workbook):
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            workbook.save(tmp.name)
            self.file_path = Path(tmp.name)
        else:
            raise TypeError("workbook must be a path or openpyxl Workbook")

    # ------------------- Helpers -------------------

    def _create_temp_range_workbook(self, sheet_name: str, cell_range: str) -> Path:
        """
        Create a temporary workbook containing only the requested range.
        """
        wb = load_workbook(self.file_path)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        ws = wb[sheet_name]

        min_col, min_row, max_col, max_row = range_boundaries(cell_range)

        tmp_wb = Workbook()
        tmp_ws = tmp_wb.active
        tmp_ws.title = sheet_name

        for r_idx, row in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row,
                                                 min_col=min_col, max_col=max_col), 1):
            for c_idx, cell in enumerate(row, 1):
                new_cell = tmp_ws.cell(row=r_idx, column=c_idx, value=cell.value)
                if cell.has_style:
                    new_cell._style = cell._style

        tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp_wb.save(tmp_file.name)
        return Path(tmp_file.name)

    def _get_chart_ranges(self, sheet_name: str) -> Dict[str, List[Tuple[int, int, int, int]]]:
        """
        Return the ranges used by charts on a sheet.

        Returns dict: chart title -> list of (min_col, min_row, max_col, max_row)
        """
        wb = load_workbook(self.file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        ws = wb[sheet_name]

        chart_ranges: Dict[str, List[Tuple[int, int, int, int]]] = {}

        for chart in ws._charts:
            ranges = []
            for s in chart.series:
                # Values
                if s.values:
                    addr = s.values.split('!')[-1].replace('$', '')
                    min_col, min_row, max_col, max_row = range_boundaries(addr)
                    ranges.append((min_col, min_row, max_col, max_row))
                # Categories
                if s.categories:
                    addr = s.categories.split('!')[-1].replace('$', '')
                    min_col, min_row, max_col, max_row = range_boundaries(addr)
                    ranges.append((min_col, min_row, max_col, max_row))
            chart_ranges[chart.title or "Unnamed Chart"] = ranges

        return chart_ranges

    def _get_union_range(self, ranges: List[Tuple[int, int, int, int]]) -> str:
        """
        Given a list of (min_col, min_row, max_col, max_row), return a single Excel range string.
        """
        if not ranges:
            raise ValueError("No ranges provided")
        min_col = min(r[0] for r in ranges)
        min_row = min(r[1] for r in ranges)
        max_col = max(r[2] for r in ranges)
        max_row = max(r[3] for r in ranges)
        return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

    # ------------------- Range Exports -------------------

    def export_range_to_pdf(self, sheet_name: str, cell_range: str, output_path: str):
        tmp_file = self._create_temp_range_workbook(sheet_name, cell_range)
        output_path = Path(output_path).absolute()

        subprocess.run([
            "libreoffice", "--headless", "--convert-to", "pdf",
            str(tmp_file), "--outdir", str(output_path.parent)
        ], check=True)

        generated_file = tmp_file.with_suffix(".pdf")
        generated_file.rename(output_path)

    def export_range_to_png(self, sheet_name: str, cell_range: str, output_path: str):
        tmp_file = self._create_temp_range_workbook(sheet_name, cell_range)
        tmp_pdf = tmp_file.with_suffix(".pdf")
        output_path = Path(output_path).absolute()

        subprocess.run([
            "libreoffice", "--headless", "--convert-to", "pdf",
            str(tmp_file), "--outdir", str(tmp_file.parent)
        ], check=True)

        subprocess.run([
            "pdftoppm", "-png", "-singlefile", str(tmp_pdf),
            str(output_path.with_suffix(''))
        ], check=True)

    # ------------------- Sheet / Chart Exports -------------------

    def export_sheet_to_pdf(self, output_path: str):
        output_path = Path(output_path).absolute()
        subprocess.run([
            "libreoffice", "--headless", "--convert-to", "pdf",
            str(self.file_path), "--outdir", str(output_path.parent)
        ], check=True)
        generated_file = self.file_path.with_suffix(".pdf")
        generated_file.rename(output_path)

    def export_sheet_to_png(self, output_path: str):
        tmp_pdf = self.file_path.with_suffix(".pdf")
        output_path = Path(output_path).absolute()

        subprocess.run([
            "libreoffice", "--headless", "--convert-to", "pdf",
            str(self.file_path), "--outdir", str(self.file_path.parent)
        ], check=True)

        subprocess.run([
            "pdftoppm", "-png", "-singlefile", str(tmp_pdf),
            str(output_path.with_suffix(''))
        ], check=True)

    # ------------------- Chart Exports -------------------

    def export_chart_to_pdf(self, sheet_name: str, chart_name: str, output_path: str):
        chart_ranges = self._get_chart_ranges(sheet_name).get(chart_name)
        if not chart_ranges:
            raise ValueError(f"Chart '{chart_name}' not found in sheet '{sheet_name}'")
        cell_range = self._get_union_range(chart_ranges)
        self.export_range_to_pdf(sheet_name, cell_range, output_path)

    def export_chart_to_png(self, sheet_name: str, chart_name: str, output_path: str):
        chart_ranges = self._get_chart_ranges(sheet_name).get(chart_name)
        if not chart_ranges:
            raise ValueError(f"Chart '{chart_name}' not found in sheet '{sheet_name}'")
        cell_range = self._get_union_range(chart_ranges)
        self.export_range_to_png(sheet_name, cell_range, output_path)
